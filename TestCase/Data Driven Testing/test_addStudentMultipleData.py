import requests
import json
import openpyxl

# data driven testing means executing multiple data on same api at a time

def test_studentmultipledata():
    # reading studentdata.json file
    f = open("studentdata.json", "r")
    # parsing it into json
    json_data = json.loads(f.read())

    # reading xml file using openpyxl
    xmlData = openpyxl.load_workbook("studentmultipledata.xlsx")
    # print(xmlData)
    xmlSheet = xmlData["Sheet1"]
    # print(xmlSheet)
    maxrows = xmlSheet.max_row
    # print(maxrows)

    for i in range(2,maxrows+1):
        cell_first_name = xmlSheet.cell(row=i, column=1)
        cell_middle_name = xmlSheet.cell(row=i, column=2)
        cell_last_name = xmlSheet.cell(row=i, column=3)
        cell_date_of_birth =  xmlSheet.cell(row=i, column=4)

        
        json_data["first_name"] = cell_first_name.value
        json_data["middle_name"] = cell_middle_name.value
        json_data["last_name"] = cell_last_name.value
        json_data["date_of_birth"] = cell_date_of_birth.value


        response = requests.post("https://thetestingworldapi.com/api/studentsDetails", data=json_data)
        print(response.text)
        assert response.status_code == 201