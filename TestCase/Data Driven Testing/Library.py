import openpyxl


class Comman:

    def __init__(self, FilePath, SheetName):
        global wk, sheet
        wk = openpyxl.load_workbook(FilePath)
        sheet = wk[SheetName]

    def fetch_row_count(self):
        return sheet.max_row
    
    def fetch_col_count(self):
        return sheet.max_column
    
    def fetch_key_name(self):
        c = sheet.max_column
        l1 = []
        for i in range(1,c+1):
            cell = sheet.cell(row=1, column=i)
            l1.insert(i-1, cell.value)    
        return l1

    def update_json_data(self, rowNumber, jsonRequest, keyList):
        c = sheet.max_column
        for i in range(1, c+1):
            cell = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value
        return jsonRequest
